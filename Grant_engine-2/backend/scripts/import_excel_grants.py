"""Import manually curated grants from draft/*.xlsx into grants_raw MongoDB.

All grants in the Excel files are manually verified by the AltCarbon team —
they skip the scout pipeline and go directly into grants_raw as 'processed: False'
so the analyst picks them up on the next run.

Usage:
    python backend/scripts/import_excel_grants.py
"""
from __future__ import annotations
import asyncio
import hashlib
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
logger = logging.getLogger("import_excel")

# ── Path setup ────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

DRAFT_DIR = PROJECT_ROOT / "draft"

# ── Currency normalisation (reuse analyst's table) ────────────────────────────
_USD_RATES = {
    "USD": 1.0, "INR": 83.5, "EUR": 0.92, "GBP": 0.79,
    "CAD": 1.36, "AUD": 1.53, "SGD": 1.34, "JPY": 149.0,
}

_AMOUNT_PATTERNS = [
    # USD / EUR / GBP
    (re.compile(r"(?:USD|US\$|\$)\s*([\d,]+(?:\.\d+)?)\s*(?:million|M\b)?", re.I), "USD"),
    (re.compile(r"(?:EUR|€)\s*([\d,]+(?:\.\d+)?)\s*(?:million|M\b)?", re.I), "EUR"),
    (re.compile(r"(?:GBP|£)\s*([\d,]+(?:\.\d+)?)\s*(?:million|M\b)?", re.I), "GBP"),
    # INR lakh / crore
    (re.compile(r"(?:INR|Rs\.?|₹)\s*([\d,]+(?:\.\d+)?)\s*(?:crore|cr\.?)\b", re.I), "INR_CR"),
    (re.compile(r"(?:INR|Rs\.?|₹)\s*([\d,]+(?:\.\d+)?)\s*(?:lakh|lac)\b", re.I), "INR_L"),
    (re.compile(r"([\d,]+(?:\.\d+)?)\s*(?:crore|cr\.?)\s*(?:INR|Rs\.?|₹)?", re.I), "INR_CR"),
    (re.compile(r"([\d,]+(?:\.\d+)?)\s*(?:lakh|lac)\s*(?:INR|Rs\.?|₹)?", re.I), "INR_L"),
    # plain numbers with M/K suffix
    (re.compile(r"(?:up to\s+)?\$([\d,]+(?:\.\d+)?)\s*[Mm]\b", re.I), "USD_M"),
    (re.compile(r"(?:up to\s+)?\$([\d,]+(?:\.\d+)?)[Kk]\b", re.I), "USD_K"),
]


def _parse_amount(raw: str) -> tuple[Optional[float], str]:
    """Return (max_funding_usd, currency) from a raw amount string."""
    if not raw:
        return None, "USD"
    raw = raw.replace(",", "")

    for pat, cur in _AMOUNT_PATTERNS:
        m = pat.search(raw)
        if not m:
            continue
        val = float(m.group(1).replace(",", ""))
        multiplier = 1
        if "million" in m.group(0).lower() or re.search(r"\bM\b", m.group(0)):
            multiplier = 1_000_000
        if cur == "USD_M":
            return val * 1_000_000, "USD"
        if cur == "USD_K":
            return val * 1_000, "USD"
        if cur == "INR_CR":
            inr = val * 10_000_000
            return inr, "INR"
        if cur == "INR_L":
            inr = val * 100_000
            return inr, "INR"
        usd = val * multiplier / _USD_RATES.get(cur, 1.0)
        return round(usd, 2), cur

    return None, "USD"


def _url_hash(url: str) -> str:
    return hashlib.md5(url.strip().lower().encode()).hexdigest()


def _content_hash(title: str, funder: str) -> str:
    def norm(s: str) -> str:
        s = s.lower().strip()
        s = re.sub(r"\s+", " ", s)
        s = re.sub(r"[^\w\s]", "", s)
        return s
    return hashlib.md5(f"{norm(title)}|{norm(funder)}".encode()).hexdigest()


def _parse_deadline(raw: str) -> Optional[str]:
    """Normalise deadline string from Excel to YYYY-MM-DD or human-readable."""
    if not raw or raw.lower() in ("", "none", "nan", "rolling", "ongoing", "tbd"):
        return "Rolling" if raw and raw.lower() in ("rolling", "ongoing") else None
    # Already a datetime string from openpyxl
    if "00:00:00" in raw:
        try:
            return raw.split(" ")[0]
        except Exception:
            pass
    return raw.strip() or None


def read_all_grants(draft_dir: Path) -> List[Dict]:
    """Read all Excel sheets and return a list of normalised grant dicts."""
    all_rows: List[Dict] = []

    for fname in sorted(os.listdir(draft_dir)):
        if not fname.endswith(".xlsx"):
            continue
        path = draft_dir / fname
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            logger.warning("Could not open %s: %s", fname, e)
            continue

        for shname in wb.sheetnames:
            ws = wb[shname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(h).strip() if h else "" for h in rows[0]]

            for row in rows[1:]:
                if not any(row):
                    continue
                d: Dict = {}
                for i, val in enumerate(row):
                    if i < len(header) and header[i]:
                        d[header[i]] = str(val).strip() if val is not None else ""
                if not (d.get("Title") or d.get("Program Name")):
                    continue
                d["_source_file"] = fname
                all_rows.append(d)

    # Deduplicate: key on (Program Name + Sponsor) — last version wins
    seen: Dict[str, Dict] = {}
    for g in all_rows:
        key = (g.get("Program Name", "") + "|" + g.get("Sponsor / Organization", "")).lower().strip()[:100]
        seen[key] = g

    return list(seen.values())


def excel_to_raw_doc(g: Dict) -> Optional[Dict]:
    """Convert one Excel row to a grants_raw document. Returns None if URL missing."""
    # Pick the best URL
    source_url = (
        g.get("Source URL", "").strip()
        or g.get("Link (Application)", "").strip()
    )
    # Skip mailto: and empty
    if not source_url or source_url.startswith("mailto:"):
        source_url = ""

    app_url = g.get("Link (Application)", "").strip()
    if app_url.startswith("mailto:"):
        app_url = ""

    canonical_url = source_url or app_url
    if not canonical_url:
        logger.debug("No URL for: %s — skipping", g.get("Program Name", "?")[:60])
        return None

    title = g.get("Program Name") or g.get("Title") or ""
    funder = g.get("Sponsor / Organization", "")
    themes_raw = g.get("Themes", "")
    geo = g.get("Geography", "")
    eligibility = g.get("Eligibility (short)", "")
    amount_raw = g.get("Amount", "")
    deadline_raw = _parse_deadline(g.get("Deadline", ""))
    notes = g.get("Notes", "")

    max_funding, currency = _parse_amount(amount_raw)

    # Detect themes
    themes_detected = []
    combined = (themes_raw + " " + title + " " + notes).lower()
    if any(k in combined for k in ["climate", "carbon", "cdr", "mrv", "net zero", "decarboni",
                                    "cleantech", "clean energy", "renewable", "hydrogen",
                                    "biochar", "enhanced weathering", "ocean alkalinity"]):
        themes_detected.append("climatetech")
    if any(k in combined for k in ["agri", "soil", "farm", "food", "crop", "regenerative",
                                    "biomanufacturing", "land use"]):
        themes_detected.append("agritech")
    if any(k in combined for k in ["ai", "machine learning", "deep learning", "ml", "data science",
                                    "neural", "predictive"]):
        themes_detected.append("ai_for_sciences")
    if any(k in combined for k in ["earth science", "remote sensing", "satellite", "geospatial",
                                    "lidar", "space", "observation", "isro", "esa incubed",
                                    "earth observation"]):
        themes_detected.append("applied_earth_sciences")
    if any(k in combined for k in ["social impact", "community", "rural", "livelihood",
                                    "inclusive", "gender", "equity", "marginalized", "youth"]):
        themes_detected.append("social_impact")

    raw_doc = {
        "title": title,
        "grant_name": title,
        "funder": funder,
        "url": canonical_url,
        "url_hash": _url_hash(canonical_url),
        "content_hash": _content_hash(title, funder),
        "source_url": source_url or canonical_url,
        "application_url": app_url or canonical_url,
        "geography": geo,
        "amount": amount_raw,
        "max_funding": max_funding,
        "max_funding_usd": max_funding,
        "currency": currency,
        "deadline": deadline_raw,
        "eligibility": eligibility,
        "themes_detected": themes_detected,
        "themes_text": themes_raw,
        "notes": notes,
        "raw_content": f"{title}\n{funder}\n{geo}\n{eligibility}\n{amount_raw}\n{notes}\n{themes_raw}",
        "source": "excel_import",
        "grant_type": g.get("Grant Type (startup/research/both)", "grant"),
        "processed": False,
        "scraped_at": datetime.now(timezone.utc).isoformat(),
        "last_verified_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "_imported_from": g.get("_source_file", ""),
    }
    return raw_doc


async def import_grants():
    """Read Excel files, convert, and upsert into grants_raw."""
    from backend.config.settings import get_settings
    from backend.db.mongo import grants_raw, get_db

    s = get_settings()

    # Initialise MongoDB (warm up the singleton)
    from backend.db.mongo import get_db
    get_db()  # triggers client creation via get_settings()

    excel_grants = read_all_grants(DRAFT_DIR)
    logger.info("Read %d unique grants from Excel files", len(excel_grants))

    col = grants_raw()

    # Load existing URL hashes to avoid duplication
    existing_hashes: set = set()
    async for doc in col.find({}, {"url_hash": 1}):
        if doc.get("url_hash"):
            existing_hashes.add(doc["url_hash"])

    imported = 0
    skipped_dup = 0
    skipped_no_url = 0

    for g in excel_grants:
        raw = excel_to_raw_doc(g)
        if raw is None:
            skipped_no_url += 1
            continue

        if raw["url_hash"] in existing_hashes:
            logger.debug("Already in DB: %s", raw["title"][:60])
            skipped_dup += 1
            continue

        try:
            from pymongo.errors import DuplicateKeyError
            await col.update_one(
                {"url_hash": raw["url_hash"]},
                {"$setOnInsert": raw},
                upsert=True,
            )
            existing_hashes.add(raw["url_hash"])
            imported += 1
            logger.info("Imported: %s — %s", raw["funder"][:40], raw["title"][:60])
        except DuplicateKeyError:
            skipped_dup += 1
        except Exception as e:
            logger.warning("Failed to import %s: %s", raw.get("title", "?")[:60], e)

    logger.info(
        "\n=== Import complete ===\n"
        "  Imported : %d\n"
        "  Skipped (dup)   : %d\n"
        "  Skipped (no URL): %d",
        imported, skipped_dup, skipped_no_url,
    )

    return imported


if __name__ == "__main__":
    count = asyncio.run(import_grants())
    print(f"\n✅ {count} grants imported into grants_raw — run the analyst to score them.")
