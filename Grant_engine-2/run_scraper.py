"""
AltCarbon Grant Scraper — standalone runner.

Runs Tavily + Exa searches → Jina content fetch → Claude Haiku extraction
→ Claude Sonnet scoring → MongoDB save → Excel export.

Usage:
    python3 run_scraper.py
    python3 run_scraper.py --queries 10   # limit search queries
    python3 run_scraper.py --no-db        # skip MongoDB, Excel only
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse, urlencode, parse_qs, urlunparse

# ── Load .env ─────────────────────────────────────────────────────────────────
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

import httpx

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("scraper")

# ── Config from env ────────────────────────────────────────────────────────────
TAVILY_KEY   = os.environ["TAVILY_API_KEY"]
EXA_KEY      = os.environ["EXA_API_KEY"]
JINA_KEY     = os.environ.get("JINA_API_KEY", "")
AI_GW_URL    = os.environ.get("AI_GATEWAY_URL", "https://ai-gateway.vercel.sh/v1")
AI_GW_KEY    = os.environ.get("AI_GATEWAY_API_KEY", "")
MONGODB_URI  = os.environ.get("MONGODB_URI", "")

HAIKU  = "anthropic/claude-haiku-4-5-20251001"
SONNET = "anthropic/claude-sonnet-4-6"

# ── Search queries ─────────────────────────────────────────────────────────────
TAVILY_QUERIES = [
    "climatetech startup grant open call 2026",
    "carbon removal CDR MRV startup funding 2026",
    "net zero decarbonisation grant program 2026",
    "climate innovation fund open call for proposals 2026",
    "agritech soil carbon grant program 2026",
    "regenerative agriculture funding open call 2026",
    "AI for climate science research grant 2026",
    "machine learning earth observation grant 2026",
    "applied earth sciences remote sensing grant 2026",
    "startup India grant climatetech 2026",
    "BIRAC ANRF DST grant open call 2026",
    "Google.org Impact Challenge climate 2026",
    "XPRIZE carbon removal challenge 2026",
    "Microsoft Climate Innovation Fund 2026",
    "World Bank IFC grant facility climate startups 2026",
    "Green Climate Fund GCF readiness grant 2026",
    "Bezos Earth Fund grant open call 2026",
    "EU Horizon EIC Accelerator climate deep tech 2026",
    "UKRI Innovate UK sustainability net zero funding 2026",
    "social impact climate startup funding 2026",
]

EXA_QUERIES = [
    "grant funding for startups measuring carbon removal and MRV verification",
    "funding for AI-powered environmental monitoring and earth observation tools",
    "grants for climate technology companies in India or globally open call",
    "philanthropic funding for soil carbon sequestration technology startups",
    "accelerator program deep tech climate startups equity-free funding 2026",
    "government grant program cleantech net zero startups open applications",
    "Bezos Earth Fund open grant call climate technology 2026",
    "BIRAC DBT biotechnology grant agritech climate India 2026",
    "EU Horizon Europe climate innovation grant open call",
    "USAID climate innovation grant program 2026",
]

# ── LLM helpers ───────────────────────────────────────────────────────────────
_openai_client = None

def _get_llm():
    global _openai_client
    if _openai_client is None:
        from openai import AsyncOpenAI
        _openai_client = AsyncOpenAI(api_key=AI_GW_KEY, base_url=AI_GW_URL)
    return _openai_client

async def chat(prompt: str, model: str = HAIKU, max_tokens: int = 800) -> str:
    client = _get_llm()
    r = await client.chat.completions.create(
        model=model,
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    )
    return r.choices[0].message.content or ""


# ── URL helpers ────────────────────────────────────────────────────────────────
_STRIP_PARAMS = {"utm_source","utm_medium","utm_campaign","utm_content",
                 "utm_term","fbclid","gclid","ref","source"}

def _url_hash(url: str) -> str:
    return hashlib.md5(url.strip().lower().encode()).hexdigest()

def _normalised_url_hash(url: str) -> str:
    try:
        p = urlparse(url.strip().lower())
        host = p.netloc.replace("www.", "")
        qs = {k: v for k, v in parse_qs(p.query).items() if k not in _STRIP_PARAMS}
        clean = urlunparse((p.scheme, host, p.path.rstrip("/"), "", urlencode(qs, doseq=True), ""))
        return hashlib.md5(clean.encode()).hexdigest()
    except Exception:
        return _url_hash(url)

def _content_hash(title: str, funder: str) -> str:
    key = f"{title.lower().strip()}|{funder.lower().strip()}"
    return hashlib.md5(key.encode()).hexdigest()

_JUNK_PATTERNS = re.compile(
    r"(news\.ycombinator|twitter\.com|linkedin\.com/feed|reddit\.com|"
    r"facebook\.com|instagram\.com|youtube\.com/watch|wikipedia\.org|"
    r"/news/|/press-release|/blog/|/article/)",
    re.I,
)
_GRANT_KEYWORDS = re.compile(
    r"\b(grant|funding|award|call for proposal|open call|apply|application|"
    r"accelerator|fellowship|prize|challenge|RFP|CFP|co.fund|programme|program)\b",
    re.I,
)

def _is_quality(url: str, content: str) -> bool:
    if _JUNK_PATTERNS.search(url):
        return False
    if len(content) < 200:
        return False
    if not _GRANT_KEYWORDS.search(content):
        return False
    return True


# ── Search ────────────────────────────────────────────────────────────────────
async def tavily_search(query: str, max_results: int = 8) -> List[Dict]:
    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.post(
            "https://api.tavily.com/search",
            json={
                "api_key": TAVILY_KEY,
                "query": query,
                "max_results": max_results,
                "search_depth": "advanced",
                "include_raw_content": True,
            },
        )
        if r.status_code != 200:
            return []
        data = r.json()
        results = []
        for item in data.get("results", []):
            results.append({
                "title":       item.get("title", ""),
                "url":         item.get("url", ""),
                "raw_content": item.get("raw_content") or item.get("content", ""),
                "source":      "tavily",
            })
        return results

async def exa_search(query: str, max_results: int = 6) -> List[Dict]:
    async with httpx.AsyncClient(timeout=20.0) as client:
        r = await client.post(
            "https://api.exa.ai/search",
            headers={"x-api-key": EXA_KEY, "Content-Type": "application/json"},
            json={
                "query": query,
                "numResults": max_results,
                "useAutoprompt": True,
                "contents": {"text": {"maxCharacters": 4000}},
            },
        )
        if r.status_code != 200:
            return []
        data = r.json()
        results = []
        for item in data.get("results", []):
            results.append({
                "title":       item.get("title", ""),
                "url":         item.get("url", ""),
                "raw_content": (item.get("text") or "")[:4000],
                "source":      "exa",
            })
        return results

async def jina_fetch(url: str) -> str:
    """Fetch clean text from a URL using Jina reader."""
    headers = {"Accept": "application/json"}
    if JINA_KEY:
        headers["Authorization"] = f"Bearer {JINA_KEY}"
    try:
        async with httpx.AsyncClient(timeout=20.0) as client:
            r = await client.get(
                f"https://r.jina.ai/{url}",
                headers=headers,
            )
            if r.status_code == 200:
                data = r.json() if r.headers.get("content-type","").startswith("application/json") else {}
                return data.get("data", {}).get("content", r.text)[:6000]
    except Exception:
        pass
    return ""


# ── Field extraction ──────────────────────────────────────────────────────────
EXTRACT_PROMPT = """Extract structured grant information from the text below.
Return ONLY valid JSON with these fields (use null if unknown):
{{
  "grant_name": "exact official grant/program name",
  "funder": "organization offering the grant",
  "grant_type": "grant|prize|challenge|accelerator|fellowship|contract|loan|equity",
  "geography": "countries/regions eligible (e.g. India, Global, EU)",
  "amount": "human-readable funding amount e.g. Up to $500,000",
  "max_funding_usd": <number or null>,
  "deadline": "deadline date string or null",
  "eligibility": "1-2 sentence eligibility summary",
  "application_url": "direct application URL or null",
  "themes": ["list","of","relevant","themes"]
}}

URL: {url}
Title: {title}
Content: {content}"""

async def extract_fields(item: Dict) -> Dict:
    content = (item.get("raw_content") or "")[:5000]
    if len(content) < 150:
        return {}
    prompt = EXTRACT_PROMPT.format(
        url=item.get("url",""),
        title=item.get("title",""),
        content=content,
    )
    try:
        raw = await chat(prompt, model=HAIKU, max_tokens=700)
        # Strip markdown code fences
        raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`")
        j = json.loads(raw)
        return j if isinstance(j, dict) else {}
    except Exception as e:
        log.debug("Extract failed for %s: %s", item.get("url",""), e)
        return {}


# ── Scoring ────────────────────────────────────────────────────────────────────
SCORE_PROMPT = """You are a grant analyst for AltCarbon, a climate technology startup in India focused on:
1. Climatetech — carbon removal, MRV, net-zero
2. Agritech — soil carbon, precision agriculture
3. AI for Sciences — AI applied to environmental and scientific problems
4. Applied Earth Sciences — remote sensing, satellite, geospatial
5. Social Impact — inclusive climate, rural communities

Evaluate this grant for AltCarbon and return ONLY valid JSON:

Grant: {title}
Funder: {funder}
Type: {grant_type}
Amount: {amount}
Geography: {geography}
Deadline: {deadline}
Eligibility: {eligibility}
Content snippet: {content}

{{
  "scores": {{
    "theme_alignment": <1-10>,
    "eligibility_confidence": <1-10>,
    "funding_amount": <1-10>,
    "deadline_urgency": <1-10>,
    "geography_fit": <1-10>,
    "competition_level": <1-10>
  }},
  "weighted_total": <float 1-10>,
  "recommended_action": "pursue|watch|auto_pass",
  "rationale": "2-3 sentence reason why AltCarbon should apply citing specific mission fit",
  "reasoning": "brief overall reasoning",
  "evidence_found": ["specific match 1", "specific match 2"],
  "evidence_gaps": ["gap 1"],
  "red_flags": []
}}

Scoring guide:
- weighted_total > 6.5 = pursue, > 5.0 = watch, else auto_pass
- geography_fit: India=10, Global=8, Unclear=5, Excludes India=0
- funding_amount: >$100k=10, >$50k=8, >$20k=6, >$3k=4"""

WEIGHTS = {
    "theme_alignment": 0.25,
    "eligibility_confidence": 0.20,
    "funding_amount": 0.20,
    "deadline_urgency": 0.15,
    "geography_fit": 0.10,
    "competition_level": 0.10,
}

async def score_grant(grant: Dict) -> Dict:
    content = (grant.get("raw_content") or "")[:3000]
    prompt = SCORE_PROMPT.format(
        title=grant.get("grant_name") or grant.get("title",""),
        funder=grant.get("funder","Unknown"),
        grant_type=grant.get("grant_type","grant"),
        amount=grant.get("amount","Unknown"),
        geography=grant.get("geography","Unknown"),
        deadline=grant.get("deadline","Unknown"),
        eligibility=grant.get("eligibility",""),
        content=content,
    )
    try:
        raw = await chat(prompt, model=SONNET, max_tokens=1000)
        raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`")
        j = json.loads(raw)
        if not isinstance(j, dict):
            return {"weighted_total": 0, "recommended_action": "auto_pass", "rationale": ""}
        # Recalculate weighted_total from individual scores for consistency
        scores = j.get("scores", {})
        if scores:
            total = sum(scores.get(k, 5) * w for k, w in WEIGHTS.items())
            j["weighted_total"] = round(total, 2)
        return j
    except Exception as e:
        log.debug("Score failed for %s: %s", grant.get("title",""), e)
        return {"weighted_total": 0, "recommended_action": "auto_pass", "rationale": ""}


# ── MongoDB save ───────────────────────────────────────────────────────────────
def save_to_mongo(grants: List[Dict]) -> int:
    if not MONGODB_URI:
        log.warning("No MONGODB_URI — skipping database save")
        return 0
    from pymongo import MongoClient, UpdateOne
    client = MongoClient(MONGODB_URI)
    db = client["altcarbon_grants"]
    col = db["grants_scored"]
    ops = []
    for g in grants:
        doc = {k: v for k, v in g.items() if k != "raw_content"}
        doc["status"] = "triage" if g.get("recommended_action") in ("pursue","watch") else "auto_pass"
        doc["scored_at"] = datetime.now(timezone.utc).isoformat()
        ops.append(UpdateOne(
            {"url_hash": g["url_hash"]},
            {"$set": doc},
            upsert=True,
        ))
    if ops:
        result = col.bulk_write(ops)
        return result.upserted_count + result.modified_count
    return 0


# ── Excel export ───────────────────────────────────────────────────────────────
def export_excel(grants: List[Dict], path: str):
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Grants (Triage + Pursue) ────────────────────────────────
    ws = wb.active
    ws.title = "AltCarbon Grants Tracker"

    # Colour palette
    HEADER_BG   = "1A3A2A"   # dark green
    HEADER_FG   = "FFFFFF"
    PURSUE_BG   = "D4EDDA"   # light green
    WATCH_BG    = "FFF3CD"   # light yellow
    PASS_BG     = "F8D7DA"   # light red
    ALT_ROW_BG  = "F5F9F7"   # very light green-white
    BORDER_COL  = "CCCCCC"

    thin = Side(style="thin", color=BORDER_COL)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column definitions: (header, field, width)
    COLUMNS = [
        ("Grant Name",        "grant_name",       40),
        ("Funder / Sponsor",  "funder",            28),
        ("Type",              "grant_type",        14),
        ("Geography",         "geography",         18),
        ("Ticket Size",       "amount",            18),
        ("Deadline",          "deadline",          16),
        ("AI Score",          "weighted_total",    10),
        ("Status",            "recommended_action",13),
        ("Themes",            "themes_str",        28),
        ("Rationale",         "rationale",         55),
        ("Eligibility",       "eligibility",       45),
        ("Apply Link",        "application_url",   40),
        ("Source URL",        "url",               40),
    ]

    # Header row
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Sort: pursue first, then watch, then others; within each group by score desc
    _ORDER = {"pursue": 0, "watch": 1, "auto_pass": 2}
    sorted_grants = sorted(
        grants,
        key=lambda g: (_ORDER.get(g.get("recommended_action",""), 3),
                       -float(g.get("weighted_total", 0)))
    )

    for row_idx, g in enumerate(sorted_grants, start=2):
        action = g.get("recommended_action", "auto_pass")
        score  = float(g.get("weighted_total", 0))

        # Row background
        if action == "pursue":
            row_bg = PURSUE_BG
        elif action == "watch":
            row_bg = WATCH_BG
        elif action == "auto_pass":
            row_bg = PASS_BG
        else:
            row_bg = ALT_ROW_BG if row_idx % 2 == 0 else "FFFFFF"

        fill = PatternFill("solid", fgColor=row_bg)

        # Flatten themes list
        themes = g.get("themes") or g.get("themes_detected") or []
        if isinstance(themes, list):
            g["themes_str"] = " · ".join(t.replace("_"," ").title() for t in themes)
        else:
            g["themes_str"] = str(themes)

        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            value = g.get(field, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if not isinstance(value, (int, float)) else value)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)

            # Score column: colour-code the number
            if field == "weighted_total":
                if score >= 7.5:
                    cell.font = Font(bold=True, color="155724", name="Calibri", size=10)
                elif score >= 6.5:
                    cell.font = Font(bold=True, color="856404", name="Calibri", size=10)
                elif score >= 5.0:
                    cell.font = Font(bold=True, color="7A3300", name="Calibri", size=10)
                else:
                    cell.font = Font(color="721C24", name="Calibri", size=10)

            # Status column
            if field == "recommended_action":
                label = {"pursue": "PURSUE", "watch": "WATCH", "auto_pass": "PASS"}.get(str(value), str(value).upper())
                cell.value = label
                if label == "PURSUE":
                    cell.font = Font(bold=True, color="155724", name="Calibri", size=10)
                elif label == "WATCH":
                    cell.font = Font(bold=True, color="856404", name="Calibri", size=10)
                else:
                    cell.font = Font(color="721C24", name="Calibri", size=10)

            # Hyperlinks for URL columns
            if field in ("apply_link", "application_url", "url") and value and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.font = Font(color="0563C1", underline="single", name="Calibri", size=10)
                cell.value = "Open link"

        # Row height
        ws.row_dimensions[row_idx].height = 60

    # ── Sheet 2: Pursue + Watch only (clean summary) ──────────────────────────
    ws2 = wb.create_sheet("Pursue & Watch")
    SUMMARY_COLS = [
        ("Grant Name",       "grant_name",       42),
        ("Funder",           "funder",            26),
        ("Type",             "grant_type",        14),
        ("Ticket Size",      "amount",            18),
        ("Geography",        "geography",         18),
        ("Deadline",         "deadline",          16),
        ("AI Score",         "weighted_total",    10),
        ("Rationale",        "rationale",         60),
        ("Eligibility",      "eligibility",       45),
        ("Apply Link",       "application_url",   40),
    ]

    for col_idx, (header, _, width) in enumerate(SUMMARY_COLS, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor="27574A")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    priority = [g for g in sorted_grants if g.get("recommended_action") in ("pursue","watch")]
    for row_idx, g in enumerate(priority, start=2):
        action = g.get("recommended_action","")
        row_bg = PURSUE_BG if action == "pursue" else WATCH_BG
        fill   = PatternFill("solid", fgColor=row_bg)
        score  = float(g.get("weighted_total", 0))

        themes = g.get("themes") or g.get("themes_detected") or []
        if isinstance(themes, list):
            g["themes_str"] = " · ".join(t.replace("_"," ").title() for t in themes)

        for col_idx, (_, field, _) in enumerate(SUMMARY_COLS, start=1):
            value = g.get(field, "") or ""
            cell = ws2.cell(row=row_idx, column=col_idx,
                            value=value if isinstance(value, (int,float)) else str(value))
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Calibri", size=10)

            if field == "weighted_total":
                if score >= 7.5:
                    cell.font = Font(bold=True, color="155724", name="Calibri", size=10)
                elif score >= 5.0:
                    cell.font = Font(bold=True, color="856404", name="Calibri", size=10)

            if field == "application_url" and value and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.font = Font(color="0563C1", underline="single", name="Calibri", size=10)
                cell.value = "Apply here"

        ws2.row_dimensions[row_idx].height = 55

    # ── Sheet 3: Stats summary ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Summary Stats")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 16

    def stat_row(label, value, row, bold=False):
        c1 = ws3.cell(row=row, column=1, value=label)
        c2 = ws3.cell(row=row, column=2, value=value)
        if bold:
            c1.font = Font(bold=True, name="Calibri", size=11)
            c2.font = Font(bold=True, name="Calibri", size=11)
        else:
            c1.font = c2.font = Font(name="Calibri", size=10)
        c1.border = c2.border = border

    pursue_list = [g for g in grants if g.get("recommended_action") == "pursue"]
    watch_list  = [g for g in grants if g.get("recommended_action") == "watch"]
    pass_list   = [g for g in grants if g.get("recommended_action") == "auto_pass"]
    capital     = sum((float(g.get("max_funding_usd") or 0)) for g in pursue_list + watch_list)

    ws3.cell(row=1, column=1, value="AltCarbon Grant Tracker — Run Summary").font = \
        Font(bold=True, size=13, color="1A3A2A", name="Calibri")
    ws3.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}").font = \
        Font(italic=True, color="888888", name="Calibri", size=10)

    rows_data = [
        ("Total grants found",      len(grants),        True),
        ("Pursue",                  len(pursue_list),   False),
        ("Watch",                   len(watch_list),    False),
        ("Auto-pass",               len(pass_list),     False),
        ("Capital targeted ($)",    f"${capital:,.0f}", True),
        ("Avg score (pursue+watch)",
         f"{sum(float(g.get('weighted_total',0)) for g in pursue_list+watch_list) / max(len(pursue_list+watch_list),1):.1f}",
         False),
    ]
    for i, (label, value, bold) in enumerate(rows_data, start=4):
        stat_row(label, value, i, bold)

    wb.save(path)
    log.info("Excel saved → %s", path)


# ── Main pipeline ─────────────────────────────────────────────────────────────
async def run(max_queries: int = 20, save_db: bool = True):
    log.info("=" * 60)
    log.info("AltCarbon Grant Scraper starting")
    log.info("=" * 60)

    # ── 1. Search ─────────────────────────────────────────────────────────────
    tavily_q = TAVILY_QUERIES[:max_queries]
    exa_q    = EXA_QUERIES[:max(1, max_queries // 2)]

    log.info("Running %d Tavily queries + %d Exa queries in parallel…",
             len(tavily_q), len(exa_q))

    tavily_tasks = [tavily_search(q) for q in tavily_q]
    exa_tasks    = [exa_search(q)    for q in exa_q]

    sem = asyncio.Semaphore(6)

    async def guarded(task):
        async with sem:
            return await task

    all_results = await asyncio.gather(
        *[guarded(t) for t in tavily_tasks + exa_tasks],
        return_exceptions=True,
    )

    raw_items: List[Dict] = []
    for r in all_results:
        if isinstance(r, Exception):
            continue
        raw_items.extend(r)

    log.info("Search found %d raw results", len(raw_items))

    # ── 2. Deduplicate by URL ─────────────────────────────────────────────────
    seen_url_hashes: set = set()
    unique_items: List[Dict] = []
    for item in raw_items:
        url = item.get("url","")
        if not url or not url.startswith("http"):
            continue
        h = _normalised_url_hash(url)
        if h in seen_url_hashes:
            continue
        seen_url_hashes.add(h)
        item["url_hash"] = h
        unique_items.append(item)

    log.info("After URL dedup: %d unique grants", len(unique_items))

    # ── 3. Fetch full content for items with thin content ─────────────────────
    async def maybe_fetch(item: Dict) -> Dict:
        if len(item.get("raw_content","")) < 500 and JINA_KEY:
            content = await jina_fetch(item["url"])
            if content:
                item["raw_content"] = content
        return item

    fetch_sem = asyncio.Semaphore(8)
    async def guarded_fetch(item):
        async with fetch_sem:
            return await maybe_fetch(item)

    unique_items = await asyncio.gather(*[guarded_fetch(i) for i in unique_items])
    unique_items = list(unique_items)

    # ── 4. Filter junk ────────────────────────────────────────────────────────
    quality = [i for i in unique_items
               if _is_quality(i.get("url",""), i.get("raw_content",""))]
    log.info("After quality filter: %d grants", len(quality))

    # ── 5. Extract structured fields (Haiku, parallel) ────────────────────────
    log.info("Extracting fields with Claude Haiku…")
    extract_sem = asyncio.Semaphore(5)
    async def guarded_extract(item):
        async with extract_sem:
            fields = await extract_fields(item)
            item.update(fields)
            return item

    quality = list(await asyncio.gather(*[guarded_extract(i) for i in quality]))

    # ── 6. Second dedup by content hash (same grant, different URLs) ──────────
    seen_content: set = set()
    deduped: List[Dict] = []
    for item in quality:
        title  = (item.get("grant_name") or item.get("title") or "").strip()
        funder = (item.get("funder") or "").strip()
        if title and funder:
            ch = _content_hash(title, funder)
            if ch in seen_content:
                continue
            seen_content.add(ch)
        deduped.append(item)

    log.info("After content dedup: %d grants", len(deduped))

    # ── 7. Score with Claude Sonnet (parallel, rate-limited) ─────────────────
    log.info("Scoring %d grants with Claude Sonnet…", len(deduped))
    score_sem = asyncio.Semaphore(3)
    async def guarded_score(item):
        async with score_sem:
            scoring = await score_grant(item)
            item.update(scoring)
            return item

    scored = list(await asyncio.gather(*[guarded_score(i) for i in deduped]))

    pursue = sum(1 for g in scored if g.get("recommended_action") == "pursue")
    watch  = sum(1 for g in scored if g.get("recommended_action") == "watch")
    log.info("Scored: %d pursue | %d watch | %d auto-pass",
             pursue, watch, len(scored) - pursue - watch)

    # ── 8. Save to MongoDB ────────────────────────────────────────────────────
    if save_db and MONGODB_URI:
        saved = save_to_mongo(scored)
        log.info("Saved %d records to MongoDB (grants_scored)", saved)

    # ── 9. Export Excel ────────────────────────────────────────────────────────
    ts   = datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(os.path.dirname(__file__), f"AltCarbon_Grants_Tracker_{ts}.xlsx")
    export_excel(scored, path)

    log.info("=" * 60)
    log.info("DONE — %d grants total | %d pursue | %d watch", len(scored), pursue, watch)
    log.info("Excel → %s", path)
    log.info("=" * 60)

    return path


# ── CLI entry ──────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="AltCarbon Grant Scraper")
    parser.add_argument("--queries", type=int, default=20,
                        help="Max Tavily queries to run (default: 20)")
    parser.add_argument("--no-db", action="store_true",
                        help="Skip MongoDB save, export Excel only")
    args = parser.parse_args()

    asyncio.run(run(max_queries=args.queries, save_db=not args.no_db))
